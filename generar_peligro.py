import json
import requests

from io import BytesIO
from datetime import datetime

from openpyxl import load_workbook


URL_XLSX = (
    "https://gencat.cat/medinatural/incendis/mapes/"
    "taula_muni_perill_avui.xlsx"
)

ARCHIVO_SALIDA = "peligro.json"


def descargar_xlsx():

    print("Descargando XLSX oficial...")

    respuesta = requests.get(
        URL_XLSX,
        timeout=30
    )

    print(
        "HTTP:",
        respuesta.status_code
    )

    if respuesta.status_code != 200:

        raise RuntimeError(
            "No se ha podido descargar el XLSX"
        )

    return respuesta.content


def normalizar_codigo(codigo):

    if codigo is None:
        return ""

    texto = str(codigo).strip()

    if texto.isdigit():

        texto = texto.zfill(5)

    return texto


def generar_json():

    contenido = descargar_xlsx()

    libro = load_workbook(
        filename=BytesIO(contenido),
        read_only=True,
        data_only=True
    )

    hoja = libro.active

    municipios = {}

    fecha_actualizacion = ""

    for fila in hoja.iter_rows(
        min_row=4,
        max_row=950,
        min_col=2,
        max_col=6,
        values_only=True
    ):

        codigo = fila[0]
        nombre = fila[1]
        comarca = fila[2]
        peligro = fila[3]
        fecha = fila[4]

        if nombre is None:
            continue

        codigo = normalizar_codigo(
            codigo
        )

        nombre = str(
            nombre
        ).strip()

        comarca = (
            str(comarca).strip()
            if comarca is not None
            else ""
        )

        peligro = (
            str(peligro).strip()
            if peligro is not None
            else ""
        )

        if isinstance(
            fecha,
            datetime
        ):

            fecha_texto = fecha.strftime(
                "%d/%m/%y"
            )

        elif fecha is not None:

            fecha_texto = str(
                fecha
            ).strip()

        else:

            fecha_texto = ""

        if fecha_texto:

            fecha_actualizacion = (
                fecha_texto
            )

        if not codigo:
            continue

        municipios[codigo] = {

            "municipio": nombre,

            "comarca": comarca,

            "peligro": peligro,

            "fecha": fecha_texto
        }

    libro.close()

    resultado = {

        "fecha": fecha_actualizacion,

        "municipios": municipios
    }

    with open(
        ARCHIVO_SALIDA,
        "w",
        encoding="utf-8"
    ) as archivo:

        json.dump(
            resultado,
            archivo,
            ensure_ascii=False,
            indent=2
        )

    print(
        "Municipios generados:",
        len(municipios)
    )

    print(
        "Archivo creado:",
        ARCHIVO_SALIDA
    )


if __name__ == "__main__":

    generar_json()
